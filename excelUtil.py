#encoding=utf-8

from openpyxl import Workbook
from openpyxl import load_workbook
import os
import string
import scipy.stats as stats


class ExcelUtil:
    _wb = None
    _ws = None
    _fileName = None
    _sheetName = None


    def __init__(self, fileName, sheetName):
        # excel file
        self._fileName = fileName
        self._sheetName = sheetName


    #创建文件并保存
    def createExcel(self):
        wb = Workbook()
        wb.create_sheet(self._sheetName)
        self._wb = wb
        wb.save(self._fileName)

    #创建工作簿及Sheet，并返回Excel对象
    def createExcelSheet(self):
        wb = Workbook()
        wb.create_sheet(title=self._sheetName, index=0)
        self._wb = wb
        return wb

    # 返会Sheet对象
    def getSheetByName(self):
        ws = self._wb.get_sheet_by_name(self._sheetName)
        self._ws = ws
        return ws

    def saveWbObjToExcel(self, wb):
        wb.save(self._fileName)

    def closeWbExcel(self, wb):
        wb.close()

    def wirteDBToExcelByWb(self, data, data_exp, exp, time):
        if self._ws == None:
            self.getSheetByName()
        data_list = data.split()
        data_list.append(data_exp)
        data_list.append(exp)
        data_list.append(time)
        self._ws.append(data_list)

    def listExcelsInDir(self, srcDir):
        filterStr = '.xlsx'
        srcExcelList = []
        for root, dirs, files in os.walk(srcDir, topdown=False):
            for name in files:
                if name.endswith(filterStr):
                    excelFileName = os.path.join(root, name)
                    srcExcelList.append(excelFileName)
        return srcExcelList

    def readExcelsToArray(self, excelFileNameList):
        srcArr = []
        for excelFileName in excelFileNameList:
            wb = load_workbook(excelFileName)
            sheets = wb.sheetnames
            # 读取第一个sheet页
            ws = wb[sheets[0]]
            srcArr.append(float(ws.cell(ws.max_row, 1).value))
            wb.close()
        return srcArr


    if __name__ == '__main__':

        for i in range(22,23):
            destDir = '/Users/liuyx/PycharmProjects/ga/results/gdb%s/Traditional' % i
            destList = listExcelsInDir(None, destDir)
            Tra = readExcelsToArray(None, destList)
            print(Tra)
            TraAvg = sum(Tra) / len(Tra)
            print("gdb%s-TraAvg: "%i, TraAvg)

            srcDir = '/Users/liuyx/PycharmProjects/ga/results/gdb%s/OneFAll' % i
            srcList = listExcelsInDir(None, srcDir)
            Imp = readExcelsToArray(None, srcList)
            print(Imp)
            ImpAvg = sum(Imp)/len(Imp)
            print("gdb%s-ImpAvg: "%i, ImpAvg)

            print("gdb%s: " %i, stats.ranksums(Tra,Imp))
            print("--------------------------------------")



